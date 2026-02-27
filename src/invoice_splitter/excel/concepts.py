from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from invoice_splitter.excel.writer import (
    ExcelWriteError,
    ensure_table_exists,
    find_table,
    append_rows_to_table,
    backup_excel,
    prune_backups,
)


@dataclass(frozen=True)
class VendorConcept:
    vendor_id: int
    concept: str
    is_default: bool
    active: bool
    sort_order: int


def _as_bool(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in {"1", "true", "yes", "y", "si", "sí"}


def load_vendor_concepts(
    *,
    excel_path: Path,
    table_name: str = "Vendor_concepts_table",
) -> Dict[int, List[VendorConcept]]:
    """
    Lee Vendor_concepts_table (si existe) y retorna dict vendor_id -> conceptos activos ordenados.
    IMPORTANTE: NO crea la tabla si no existe.
    """
    try:
        wb = load_workbook(excel_path, data_only=True)
    except PermissionError as e:
        raise ExcelWriteError(
            "No se puede abrir el Excel para leer conceptos. Probablemente está abierto/bloqueado."
        ) from e

    try:
        found = _find_table_strict(wb, table_name)
        if not found:
            # No existe aún -> catálogo vacío (no creamos nada)
            return {}

        ws, min_col, min_row, max_col, max_row, headers = found

        idx = {str(h): i for i, h in enumerate(headers)}
        required = {"Vendor ID", "Concept", "Is_default", "Active", "Sort_order"}
        missing = required - set(idx.keys())
        if missing:
            raise ExcelWriteError(
                f"Vendor_concepts_table no tiene columnas requeridas: {sorted(missing)}"
            )

        concepts_by_vendor: Dict[int, List[VendorConcept]] = {}

        for r in range(min_row + 1, max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
            raw_vid = row_vals[idx["Vendor ID"]]
            raw_concept = row_vals[idx["Concept"]]

            if raw_vid is None or raw_concept is None:
                continue

            try:
                vendor_id = int(str(raw_vid).strip())
            except ValueError:
                continue

            concept = str(raw_concept).strip()
            if not concept:
                continue

            is_default = _as_bool(row_vals[idx["Is_default"]])
            active = _as_bool(row_vals[idx["Active"]])
            try:
                sort_order = int(str(row_vals[idx["Sort_order"]] or "0").strip())
            except ValueError:
                sort_order = 0

            vc = VendorConcept(
                vendor_id=vendor_id,
                concept=concept,
                is_default=is_default,
                active=active,
                sort_order=sort_order,
            )
            concepts_by_vendor.setdefault(vendor_id, []).append(vc)

        # filtrar activos + ordenar por sort_order
        for vid in list(concepts_by_vendor.keys()):
            active_items = [c for c in concepts_by_vendor[vid] if c.active]
            active_items.sort(key=lambda x: (x.sort_order, x.concept.lower()))
            concepts_by_vendor[vid] = active_items

        return concepts_by_vendor

    finally:
        wb.close()


def _find_table_strict(wb, table_name: str):
    """
    Busca una tabla por nombre SIN crearla si no existe.
    Retorna: (ws, min_col, min_row, max_col, max_row, headers) o None
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if table_name in ws.tables:
            table = ws.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            headers = [ws.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]
            return ws, min_col, min_row, max_col, max_row, headers
    return None


def add_concepts_for_vendor(
    *,
    excel_path: Path,
    backup_dir: Path,
    vendor_id: int,
    concepts_to_add: List[str],
    table_name: str = "Vendor_concepts_table",
) -> None:
    """
    Inserta conceptos nuevos para un vendor:
    - Active=True, Is_default=False
    - Sort_order incremental
    - Evita duplicados (vendor_id + concept)
    IMPORTANTE: aquí SÍ creamos la tabla si no existe y guardamos el workbook.
    """
    concepts_to_add = [(c or "").strip() for c in concepts_to_add]
    concepts_to_add = [c for c in concepts_to_add if c]
    if not concepts_to_add:
        return

    backup_excel(excel_path, backup_dir)
    prune_backups(backup_dir)

    try:
        wb = load_workbook(excel_path)
    except PermissionError as e:
        raise ExcelWriteError(
            "No se puede abrir el Excel para guardar conceptos (bloqueado)."
        ) from e

    try:
        # si no existe, la creamos en hoja Config (según tu writer) y seguimos
        ws, info = find_table(
            wb, table_name
        )  # find_table puede crear en memoria [1](https://kelloggcompany-my.sharepoint.com/personal/andres_saavedra_kellogg_com1/Documents/Desktop/temp/main_window.py)

        # headers
        idx = {str(h): i for i, h in enumerate(info.headers)}
        required = {"Vendor ID", "Concept", "Is_default", "Active", "Sort_order"}
        missing = required - set(idx.keys())
        if missing:
            raise ExcelWriteError(
                f"Vendor_concepts_table no tiene columnas requeridas: {sorted(missing)}"
            )

        # existentes y max sort_order
        existing = set()
        max_order = 0

        for r in range(info.min_row + 1, info.max_row + 1):
            row_vals = [
                ws.cell(row=r, column=c).value for c in range(info.min_col, info.max_col + 1)
            ]
            raw_vid = row_vals[idx["Vendor ID"]]
            raw_concept = row_vals[idx["Concept"]]
            if raw_vid is None or raw_concept is None:
                continue
            try:
                vid = int(str(raw_vid).strip())
            except ValueError:
                continue
            concept = str(raw_concept).strip()
            if not concept:
                continue

            if vid == vendor_id:
                existing.add(concept.lower())
                try:
                    so = int(str(row_vals[idx["Sort_order"]] or "0").strip())
                    max_order = max(max_order, so)
                except ValueError:
                    pass

        order = max_order + 1
        rows = []
        for c in concepts_to_add:
            if c.lower() in existing:
                continue
            rows.append(
                {
                    "Vendor ID": vendor_id,
                    "Concept": c,
                    "Is_default": False,
                    "Active": True,
                    "Sort_order": order,
                }
            )
            existing.add(c.lower())
            order += 1

        if rows:
            # refrescar info y append
            ws, info = find_table(wb, table_name)
            append_rows_to_table(ws, info, rows)

        wb.save(excel_path)  # ✅ aquí persiste realmente en el archivo

    finally:
        wb.close()
