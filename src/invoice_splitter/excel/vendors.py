from __future__ import annotations

from dataclasses import dataclass
from typing import List, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


from pathlib import Path
from typing import Optional, Dict, Any, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from invoice_splitter.excel.writer import backup_excel, prune_backups, ExcelWriteError


@dataclass(frozen=True)
class Vendor:
    """
    Representa un vendor leído desde Vendors_table.
    vendor_id se guarda como int si es posible (para comparaciones),
    vendor_name como string (tal cual aparece en Excel).
    """

    vendor_id: int
    vendor_name: str


def load_vendors_from_table(
    excel_path: str,
    sheet_name: str = "Vendors",
    table_name: str = "Vendors_table",
) -> List[Vendor]:
    """
    Carga la lista de vendors desde una Excel Table (ListObject) llamada table_name
    dentro de la hoja sheet_name.

    Requisitos:
    - La tabla debe tener columnas "ID" y "Vendor" (encabezados).
    - No modifica el archivo Excel.
    - Devuelve lista ordenada por vendor_name.

    Manejo de errores:
    - Si el archivo está bloqueado y Windows no permite lectura, se lanza PermissionError
      con mensaje claro.
    """
    try:
        wb = load_workbook(excel_path, data_only=True, read_only=False)
    except PermissionError as e:
        raise PermissionError(
            "No se puede abrir el Excel. Probablemente está abierto en modo exclusivo "
            "o bloqueado por OneDrive/Excel. Cierra el archivo e inténtalo de nuevo."
        ) from e

    if sheet_name not in wb.sheetnames:
        raise KeyError(f"No existe la hoja '{sheet_name}' en el archivo Excel.")

    ws = wb[sheet_name]

    # openpyxl guarda tablas en ws.tables (dict nombre -> objeto Table)
    if table_name not in ws.tables:
        available = ", ".join(ws.tables.keys()) or "(ninguna)"
        raise KeyError(
            f"No existe la tabla '{table_name}' en la hoja '{sheet_name}'. "
            f"Tablas disponibles: {available}"
        )

    table = ws.tables[table_name]
    # table.ref es un rango tipo "A1:B20"
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    # Leemos encabezados (fila 1 del rango de la tabla)
    headers = []
    for col in range(min_col, max_col + 1):
        headers.append(ws.cell(row=min_row, column=col).value)

    # Normalizamos índices de columnas
    # Esperamos "ID" y "Vendor" exactamente (si difieren en Excel, ajustamos aquí)
    try:
        id_idx = headers.index("ID")
        vendor_idx = headers.index("Vendor")
    except ValueError:
        raise ValueError(
            f"Encabezados inválidos en {table_name}. "
            f"Se esperaban columnas 'ID' y 'Vendor'. Encabezados encontrados: {headers}"
        )

    vendors: List[Vendor] = []

    # Filas de datos: desde min_row + 1 hasta max_row
    for row in range(min_row + 1, max_row + 1):
        row_values = []
        for col in range(min_col, max_col + 1):
            row_values.append(ws.cell(row=row, column=col).value)

        raw_id = row_values[id_idx]
        raw_vendor = row_values[vendor_idx]

        # Saltar filas vacías
        if raw_id is None and raw_vendor is None:
            continue

        if raw_id is None or raw_vendor is None:
            # Si hay filas incompletas, mejor fallar con mensaje claro
            raise ValueError(
                f"Fila incompleta en {table_name} ({sheet_name}), fila Excel {row}. "
                f"ID={raw_id}, Vendor={raw_vendor}"
            )

        try:
            vendor_id = int(str(raw_id).strip())
        except ValueError as e:
            raise ValueError(f"ID de vendor inválido en {table_name} fila {row}: {raw_id}") from e

        vendor_name = str(raw_vendor).strip()
        vendors.append(Vendor(vendor_id=vendor_id, vendor_name=vendor_name))

    wb.close()

    # Ordenamos alfabéticamente para el combobox
    vendors.sort(key=lambda v: v.vendor_name.lower())
    return vendors


def add_vendor_to_table(
    *,
    excel_path: Path,
    sheet_name: str,
    table_name: str,
    vendor_id: int,
    vendor_name: str,
    backup_dir: Path,
    keep_last_n: int = 30,
    keep_days: int = 30,
) -> None:
    """
    Inserta un vendor nuevo en Vendors_table (columnas ID y Vendor).
    - Bloquea si el ID ya existe (la UI ya lo valida, aquí lo validamos de nuevo por seguridad).
    - Hace backup por sesión simple (crea uno nuevo por cada llamada).
    """
    # Backup (simple)
    backup_excel(excel_path, backup_dir)
    prune_backups(backup_dir, keep_last_n=keep_last_n, keep_days=keep_days)

    try:
        wb = load_workbook(excel_path)
    except PermissionError as e:
        raise ExcelWriteError(
            "No se puede abrir el Excel (probablemente está abierto o bloqueado por OneDrive/SharePoint)."
        ) from e

    try:
        if sheet_name not in wb.sheetnames:
            raise ExcelWriteError(f"No existe la hoja '{sheet_name}'.")

        ws = wb[sheet_name]
        if table_name not in ws.tables:
            raise ExcelWriteError(f"No existe la tabla '{table_name}' en la hoja '{sheet_name}'.")

        table = ws.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)

        # headers
        headers = [ws.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]
        if "ID" not in headers or "Vendor" not in headers:
            raise ExcelWriteError(
                f"Encabezados inválidos en {table_name}. Se requieren columnas 'ID' y 'Vendor'. "
                f"Encontrados: {headers}"
            )

        id_idx = headers.index("ID")
        vendor_idx = headers.index("Vendor")

        id_col = min_col + id_idx
        vendor_col = min_col + vendor_idx

        # Validación de ID duplicado
        for r in range(min_row + 1, max_row + 1):
            cell_id = ws.cell(row=r, column=id_col).value
            if cell_id is None:
                continue
            try:
                if int(str(cell_id).strip()) == vendor_id:
                    raise ExcelWriteError(f"El Vendor ID {vendor_id} ya existe en {table_name}.")
            except ValueError:
                # si hay basura en Excel, lo ignoramos aquí
                continue

        # Insertar nueva fila al final de la tabla
        new_row = max_row + 1
        ws.cell(row=new_row, column=id_col, value=vendor_id)
        ws.cell(row=new_row, column=vendor_col, value=vendor_name)

        # Expandir ref
        start_cell = f"{get_column_letter(min_col)}{min_row}"
        end_cell = f"{get_column_letter(max_col)}{new_row}"
        ws.tables[table_name].ref = f"{start_cell}:{end_cell}"

        # Guardar
        wb.save(excel_path)

    finally:
        wb.close()


# Pequeña prueba manual: permite ejecutar este archivo directamente
if __name__ == "__main__":
    from invoice_splitter.config import get_settings

    s = get_settings()
    vendors = load_vendors_from_table(
        excel_path=str(s.excel_path),
        sheet_name=s.vendors_sheet,
        table_name=s.vendors_table,
    )
    print(f"Vendors cargados: {len(vendors)}")
    print("Primeros 10:")
    for v in vendors[:10]:
        print(v)
