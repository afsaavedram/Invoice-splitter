from __future__ import annotations

import logging
import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import Table

logger = logging.getLogger("invoice_splitter")


@dataclass(frozen=True)
class TableInfo:
    sheet_name: str
    table_name: str
    min_col: int
    min_row: int
    max_col: int
    max_row: int
    headers: List[str]


class ExcelWriteError(RuntimeError):
    """Errores controlados al escribir Excel (archivo bloqueado, tabla no encontrada, etc.)."""


# -----------------------
# Excel sheet naming helpers (31 chars + invalid chars)
# -----------------------

_INVALID_SHEET_CHARS = r"[:\\/?*\[\]]"  # Excel no permite : \ / ? * [ ]
_MAX_SHEET_LEN = 31  # Límite Excel 31 caracteres


def sanitize_sheet_name(name: str) -> str:
    """
    Normaliza un nombre de hoja Excel:
    - Elimina caracteres inválidos (: \ / ? * [ ])
    - Recorta a 31 caracteres
    - Evita vacío
    """
    import re

    s = (name or "").strip()
    s = re.sub(_INVALID_SHEET_CHARS, "", s)
    if not s:
        s = "Sheet"
    s = s[:_MAX_SHEET_LEN]
    return s


def unique_sheet_name(wb, base: str) -> str:
    """
    Asegura unicidad del nombre de hoja. Si existe, usa sufijos _2, _3, ...
    Manteniendo longitud <= 31.
    """
    base = sanitize_sheet_name(base)
    if base not in wb.sheetnames:
        return base

    n = 2
    while True:
        suffix = f"_{n}"
        max_base = _MAX_SHEET_LEN - len(suffix)
        candidate = (base[:max_base] + suffix)[:_MAX_SHEET_LEN]
        if candidate not in wb.sheetnames:
            return candidate
        n += 1


# -----------------------
# Table schemas (headers) for auto-creation
# -----------------------

BASE_HEADERS = [
    "Date",
    "Bill number",
    "ID",
    "Vendor",
    "Service/ concept",
    "CC",
    "GL account",
    "Subtotal assigned by CC",
    "% IVA",
    "IVA assigned by CC",
    "Total assigned by CC",
]  # Coinciden con lo que ya escribe el writer y la UI [1](https://support.microsoft.com/en-us/office/excel-doesn-t-fully-support-some-special-characters-in-the-filename-or-folder-path-20728217-f08a-4d63-a741-821a14cec380)[1](https://support.microsoft.com/en-us/office/excel-doesn-t-fully-support-some-special-characters-in-the-filename-or-folder-path-20728217-f08a-4d63-a741-821a14cec380)

EXTRA_HEADERS_BY_TABLE = {
    "Cirion_table": [
        "Bandwidth (MBPS)"
    ],  # [1](https://support.microsoft.com/en-us/office/excel-doesn-t-fully-support-some-special-characters-in-the-filename-or-folder-path-20728217-f08a-4d63-a741-821a14cec380)[1](https://support.microsoft.com/en-us/office/excel-doesn-t-fully-support-some-special-characters-in-the-filename-or-folder-path-20728217-f08a-4d63-a741-821a14cec380)
    "Claro_siptrunk_table": [
        "Bandwidth (MBPS)",
        "Troncal SIP (channels)",
    ],  # [6](https://stackoverflow.com/questions/63182578/renaming-excel-sheets-name-exceeds-31-characters-error)[1](https://support.microsoft.com/en-us/office/excel-doesn-t-fully-support-some-special-characters-in-the-filename-or-folder-path-20728217-f08a-4d63-a741-821a14cec380)
    "Claro_SBC_table": [
        "Siptrunk (MBPS)",
        "Licences (Quantity)",
        "Siptrunk price",
        "Licences price",
    ],  # [6](https://stackoverflow.com/questions/63182578/renaming-excel-sheets-name-exceeds-31-characters-error)[1](https://support.microsoft.com/en-us/office/excel-doesn-t-fully-support-some-special-characters-in-the-filename-or-folder-path-20728217-f08a-4d63-a741-821a14cec380)
    "Movistar_table": [
        "Phone lines quantity"
    ],  # [6](https://stackoverflow.com/questions/63182578/renaming-excel-sheets-name-exceeds-31-characters-error)[1](https://support.microsoft.com/en-us/office/excel-doesn-t-fully-support-some-special-characters-in-the-filename-or-folder-path-20728217-f08a-4d63-a741-821a14cec380)
    "Claro_mobile_table": [
        "Phone lines quantity"
    ],  # [6](https://stackoverflow.com/questions/63182578/renaming-excel-sheets-name-exceeds-31-characters-error)[1](https://support.microsoft.com/en-us/office/excel-doesn-t-fully-support-some-special-characters-in-the-filename-or-folder-path-20728217-f08a-4d63-a741-821a14cec380)
}


def get_table_headers(table_name: str) -> list[str]:
    return BASE_HEADERS + EXTRA_HEADERS_BY_TABLE.get(table_name, [])


# -----------------------
# Backups + retención
# -----------------------
def backup_excel(excel_path: Path, backup_dir: Path) -> Path:
    """Crea backup con timestamp en backup_dir."""
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = backup_dir / f"{excel_path.stem}_backup_{ts}{excel_path.suffix}"
    shutil.copy2(excel_path, dst)
    return dst


def prune_backups(backup_dir: Path, keep_last_n: int = 30, keep_days: int = 30) -> None:
    """
    Retención combinada:
    1) Borra backups con antigüedad > keep_days
    2) Luego limita a keep_last_n backups más recientes (entre los restantes)

    Esto cumple: "conservar últimos 30 backups o 30 días (lo que pase primero)".
    """
    if not backup_dir.exists():
        return

    now = datetime.now()
    cutoff = now - timedelta(days=keep_days)

    # Backups típicos: *_backup_YYYYMMDD_HHMMSS.xlsx (o .xlsm si algún día cambias)
    backups = [
        p
        for p in backup_dir.glob("*_backup_*.*")
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"}
    ]

    # 1) por antigüedad
    for p in backups:
        try:
            mtime = datetime.fromtimestamp(p.stat().st_mtime)
            if mtime < cutoff:
                p.unlink()
                logger.info("RETENCION | eliminado por antigüedad | %s", p)
        except Exception as e:
            logger.warning("RETENCION | no se pudo eliminar %s | %s", p, e)

    # 2) por cantidad (recalcular después del borrado)
    backups = [
        p
        for p in backup_dir.glob("*_backup_*.*")
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"}
    ]
    backups.sort(key=lambda p: p.stat().st_mtime, reverse=True)

    for p in backups[keep_last_n:]:
        try:
            p.unlink()
            logger.info("RETENCION | eliminado por exceso de cantidad | %s", p)
        except Exception as e:
            logger.warning("RETENCION | no se pudo eliminar %s | %s", p, e)


# -----------------------
# Workbook helpers
# -----------------------
def open_workbook_safe(excel_path: Path):
    """Abre el workbook controlando archivo bloqueado."""
    try:
        return load_workbook(excel_path)
    except PermissionError as e:
        raise ExcelWriteError(
            "No se puede abrir el Excel. Probablemente está abierto en Excel o bloqueado por OneDrive.\n"
            "Cierra el archivo y vuelve a intentarlo."
        ) from e


def find_table(wb, table_name: str) -> Tuple[Any, TableInfo]:
    """Busca una Excel Table por nombre a través de todas las hojas.
    Si no existe, la crea (sheet==table por default).
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if table_name in ws.tables:
            table = ws.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            headers = []
            for c in range(min_col, max_col + 1):
                headers.append(ws.cell(row=min_row, column=c).value)
            return ws, TableInfo(
                sheet_name=sheet_name,
                table_name=table_name,
                min_col=min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
                headers=headers,
            )

    # Si no se encontró, crear hoja + tabla automáticamente (Fase 1)
    ws, info = ensure_table_exists(wb, table_name, sheet_name=table_name)
    logger.info("TABLA CREADA AUTOMATICAMENTE tabla=%s hoja=%s", table_name, info.sheet_name)
    return ws, info


def ensure_table_exists(
    wb, table_name: str, sheet_name: str | None = None
) -> tuple[Any, TableInfo]:
    """
    Si la tabla no existe:
    - Crea hoja (default sheet == table)
    - Escribe headers en fila 1
    - Crea Excel Table (ListObject) con ref A1:<lastcol>2 (Opción A)
      (fila 2 queda vacía sin valores; luego se llena en el primer append)
    """
    headers = get_table_headers(table_name)

    # Default: sheet == table (como acordamos)
    sheet_base = sheet_name or table_name
    sheet_base = sanitize_sheet_name(sheet_base)
    sheet_final = unique_sheet_name(wb, sheet_base)

    # Crear hoja si no existe (o usar existente si ya está)
    if sheet_final in wb.sheetnames:
        ws = wb[sheet_final]
    else:
        ws = wb.create_sheet(
            sheet_final
        )  # openpyxl create_sheet [3](https://www.splitmyinvoice.com/)

    # Escribir headers en fila 1 si está vacía (o si no hay nada)
    # (si el usuario ya creó una hoja con algo, esto no borra nada; solo escribe si A1 está vacío)
    if ws.cell(row=1, column=1).value is None:
        for idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=h)

    # Crear la tabla con ref A1:<lastcol>2 (Opción A)
    from openpyxl.utils import get_column_letter

    last_col_letter = get_column_letter(len(headers))
    ref = f"A1:{last_col_letter}2"  # header + 1 fila vacía (sin dummy data)
    table = Table(
        displayName=table_name, ref=ref
    )  # openpyxl Table [2](https://github.com/dotKz/api-mega-list)[3](https://www.splitmyinvoice.com/)
    ws.add_table(
        table
    )  # add_table [2](https://github.com/dotKz/api-mega-list)[3](https://www.splitmyinvoice.com/)

    # Construir TableInfo consistente con tu estructura
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    info = TableInfo(
        sheet_name=ws.title,
        table_name=table_name,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
        headers=headers,
    )
    return ws, info


def _header_to_col_index(info: TableInfo) -> Dict[str, int]:
    """Mapa 'header' -> columna absoluta (en la hoja)."""
    mapping: Dict[str, int] = {}
    for idx, header in enumerate(info.headers):
        mapping[str(header)] = info.min_col + idx
    return mapping


def _copy_row_style(ws, src_row: int, dst_row: int, min_col: int, max_col: int) -> None:
    """Copia estilo de una fila a otra para preservar formato visible."""
    for col in range(min_col, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst._style = copy(src._style)
            dst.number_format = src.number_format
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.alignment = copy(src.alignment)
            dst.protection = copy(src.protection)
            dst.comment = copy(src.comment) if src.comment else None


def _set_cell_formats(ws, row: int, col_map: Dict[str, int]) -> None:
    """
    Aplica formatos específicos:
    - Date: dd-mmm-yy
    - Bill number: texto
    - Números: 2 decimales
    - CC/GL: enteros
    """
    if "Date" in col_map:
        ws.cell(row=row, column=col_map["Date"]).number_format = "dd-mmm-yy"

    if "Bill number" in col_map:
        ws.cell(row=row, column=col_map["Bill number"]).number_format = "@"

    for key in ["Subtotal assigned by CC", "% IVA", "IVA assigned by CC", "Total assigned by CC"]:
        if key in col_map:
            ws.cell(row=row, column=col_map[key]).number_format = "#,##0.00"

    for key in ["CC", "GL account"]:
        if key in col_map:
            ws.cell(row=row, column=col_map[key]).number_format = "0"


# -----------------------
# Delete + Append
# -----------------------
def delete_duplicates_in_table(ws, info: TableInfo, vendor_id: int, bill_number: str) -> int:
    """
    Borra filas dentro de la tabla donde:
      (ID == vendor_id) y (Bill number == bill_number)

    Duplicado definido por ti: SOLO vendor_id + bill_number (sin fecha).
    Además, ENCOGE el table.ref para evitar huecos.
    """
    col_map = _header_to_col_index(info)

    if "ID" not in col_map or "Bill number" not in col_map:
        raise ExcelWriteError(
            f"La tabla {info.table_name} debe tener columnas 'ID' y 'Bill number' para sobrescritura."
        )

    id_col = col_map["ID"]
    bill_col = col_map["Bill number"]

    rows_to_delete: List[int] = []
    for r in range(info.min_row + 1, info.max_row + 1):
        cell_id = ws.cell(row=r, column=id_col).value
        cell_bill = ws.cell(row=r, column=bill_col).value
        if cell_id == vendor_id and str(cell_bill).strip() == bill_number:
            rows_to_delete.append(r)

    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

    deleted = len(rows_to_delete)

    # CLAVE: encoger el rango de la tabla para evitar huecos
    if deleted > 0:
        start_cell = f"{get_column_letter(info.min_col)}{info.min_row}"
        new_end_row = info.max_row - deleted
        end_cell = f"{get_column_letter(info.max_col)}{new_end_row}"
        ws.tables[info.table_name].ref = f"{start_cell}:{end_cell}"

    return deleted


def _is_row_empty(ws, row: int, min_col: int, max_col: int) -> bool:
    for col in range(min_col, max_col + 1):
        if ws.cell(row=row, column=col).value not in (None, ""):
            return False
    return True


def append_rows_to_table(ws, info: TableInfo, rows: List[Dict[str, Any]]) -> None:
    """Inserta filas al final de la tabla y expande table.ref.
    - Si la tabla recién se creó con ref ...:2 (fila 2 vacía), la primera inserción se escribe en fila 2.
    """
    col_map = _header_to_col_index(info)

    # Determinar la primera fila de datos dentro de la tabla (normalmente row 2)
    first_data_row = info.min_row + 1

    # Si la tabla fue creada con ref hasta row 2, y row 2 está vacía, empezamos ahí
    if info.max_row >= first_data_row and _is_row_empty(
        ws, first_data_row, info.min_col, info.max_col
    ):
        insert_row = first_data_row
        last_data_row = first_data_row - 1  # todavía no hay data real
    else:
        insert_row = info.max_row + 1
        last_data_row = info.max_row

    # Para tablas existentes con estilos, intentamos copiar estilo de la última fila de datos
    # Para tablas nuevas (sin estilo), simplemente no copiamos.
    def has_any_style(r: int) -> bool:
        for c in range(info.min_col, info.max_col + 1):
            if ws.cell(row=r, column=c).has_style:
                return True
        return False

    style_source_row = (
        last_data_row if last_data_row >= first_data_row and has_any_style(last_data_row) else None
    )

    for row_values in rows:
        new_row = insert_row

        # copiar estilo solo si hay fuente válida (tablas existentes)
        if style_source_row is not None:
            _copy_row_style(ws, style_source_row, new_row, info.min_col, info.max_col)

        # escribir valores
        for header, value in row_values.items():
            if header not in col_map:
                raise ExcelWriteError(
                    f"Columna '{header}' no existe en la tabla {info.table_name}. "
                    f"Headers disponibles: {info.headers}"
                )
            ws.cell(row=new_row, column=col_map[header], value=value)

        # formatos (sí importa consistencia)
        _set_cell_formats(ws, new_row, col_map)

        # avanzar punteros
        last_data_row = new_row
        style_source_row = new_row if has_any_style(new_row) else style_source_row
        insert_row = new_row + 1

    # actualizar ref de la tabla (expandir hasta last_data_row)
    start_cell = f"{get_column_letter(info.min_col)}{info.min_row}"
    end_cell = f"{get_column_letter(info.max_col)}{last_data_row if last_data_row >= first_data_row else first_data_row}"
    ws.tables[info.table_name].ref = f"{start_cell}:{end_cell}"


# -----------------------
# Transaction (Backup por sesión + retención + logging)
# -----------------------
def apply_transaction(
    excel_path: Path,
    backup_dir: Path,
    vendor_id: int,
    bill_number: str,
    table_to_rows: Dict[str, List[Dict[str, Any]]],
    backup_path: Path | None = None,
    retention_keep_last_n: int = 30,
    retention_keep_days: int = 30,
) -> Tuple[Path, Dict[str, int], bool]:
    """
    Ejecuta un guardado completo (multi-tabla) con:
    - Backup por sesión:
        - Si backup_path es None -> crea backup (solo una vez por sesión) + aplica retención
        - Si backup_path ya existe -> reutiliza, NO crea nuevo backup
    - Abre workbook 1 vez
    - Por cada tabla destino:
        - borra duplicados por (vendor_id + bill_number) SOLO en esa tabla
        - inserta filas al final
    - Guarda workbook 1 vez

    Devuelve:
      (backup_path_usado, deleted_by_table, backup_creado_esta_vez)
    """
    backup_created = False

    # Backup por sesión
    if backup_path is None or not Path(backup_path).exists():
        backup_path = backup_excel(excel_path, backup_dir)
        backup_created = True
        prune_backups(backup_dir, keep_last_n=retention_keep_last_n, keep_days=retention_keep_days)

    wb = open_workbook_safe(excel_path)
    deleted_by_table: Dict[str, int] = {}

    logger.info(
        "TRANSACCION INICIO | excel=%s | vendor_id=%s | bill=%s | backup=%s | backup_creado=%s | tablas=%s",
        excel_path,
        vendor_id,
        bill_number,
        backup_path,
        backup_created,
        list(table_to_rows.keys()),
    )

    try:
        for table_name, rows in table_to_rows.items():
            ws, info = find_table(wb, table_name)

            deleted = delete_duplicates_in_table(ws, info, vendor_id, bill_number)
            deleted_by_table[table_name] = deleted

            # refrescar info después de borrar (rango pudo cambiar)
            ws, info = find_table(wb, table_name)
            append_rows_to_table(ws, info, rows)

            logger.info(
                "TABLA ACTUALIZADA | tabla=%s | borradas=%s | insertadas=%s",
                table_name,
                deleted,
                len(rows),
            )

        try:
            wb.save(excel_path)
        except PermissionError as e:
            logger.error("ERROR GUARDADO (archivo bloqueado) | %s", e)
            raise ExcelWriteError(
                "No se pudo guardar el Excel. Probablemente está abierto/bloqueado.\n"
                "Cierra el archivo y vuelve a intentar."
            ) from e

        logger.info(
            "TRANSACCION OK | excel=%s | vendor_id=%s | bill=%s | borradas_por_tabla=%s",
            excel_path,
            vendor_id,
            bill_number,
            deleted_by_table,
        )

    except Exception as e:
        logger.exception(
            "ERROR TRANSACCION | excel=%s | vendor_id=%s | bill=%s | %s",
            excel_path,
            vendor_id,
            bill_number,
            e,
        )
        raise
    finally:
        wb.close()

    return Path(backup_path), deleted_by_table, backup_created
